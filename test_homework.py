import os.path
from pypdf import PdfReader
from openpyxl import load_workbook
from zipfile import ZipFile, ZIP_DEFLATED

CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
TMP_DIR = os.path.join(CURRENT_DIR, 'tmp')
RESOURSES_DIR = os.path.join(CURRENT_DIR, 'resourses')


def create_archive():
    if not os.path.exists('resources/'):
        os.mkdir('resources')

    source_dir = os.path.abspath('tmp')
    dir_for_archive = os.path.abspath('resources/')

    os.chdir(TMP_DIR)
    list_files = [os.path.join(source_dir, 'test.csv'),
                  os.path.join(source_dir, 'test.xlsx'),
                  os.path.join(source_dir, 'test-resume.pdf')]
    zip_name = os.path.join(dir_for_archive, 'test.zip')
    if os.path.exists(zip_name):
        os.remove(zip_name)

    with ZipFile(zip_name, 'w') as myzip:
        for file in list_files:
            myzip.write(file, arcname=os.path.basename(file))
    assert os.path.exists(zip_name) is True
    return zip_name


def test_files_in_zip():
    zip_name = create_archive()
    with ZipFile(zip_name) as zipbox:
        with zipbox.open('test-resume.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            page = reader.pages[0]
            text = page.extract_text()
            number_of_pages = len(reader.pages)
            assert number_of_pages == 2
            assert 'Фитнес для всех' in text
    with ZipFile(zip_name) as zipbox:
        with zipbox.open('test.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            cell = sheet['D1']
            result_assert = sheet.cell(row=5, column=3).value
            assert cell.value == 'Salary'
            assert result_assert == 'Jacob'
    with ZipFile(zip_name) as zipbox:
        with zipbox.open('test.csv') as csv_file:
            for line_number, string in enumerate(csv_file, start=1):
                if line_number == 3:
                    assert string == b'"John ""Da Man""",Repici,120 Jefferson St.,Riverside, NJ,08075\n'
            assert line_number == 6
